#!/usr/bin/env python3
"""
Inventario automático: por cada región, busca en una carpeta de Google Drive
el archivo de bodega MÁS RECIENTE cuyo nombre contenga el nombre de la región,
lee REFERENCIA + SALDO (+ TAT, MARCA, DETALLE) y lo envía al backend
(/api/importar/inventario-auto) que actualiza la existencia de la bodega de esa región.

Secrets requeridos (GitHub → Settings → Secrets and variables → Actions):
  GDRIVE_SA_KEY     JSON de la cuenta de servicio de Google (todo el contenido)
  GDRIVE_FOLDER_ID  ID de la carpeta de Drive donde la bodega deja los archivos
  API_BASE          https://ventamax-pro.onrender.com/api
  IMPORT_TOKEN      el mismo token configurado en Render (IMPORT_TOKEN)
  INV_REGIONES      (opcional) lista separada por comas. Default: QUINDIO,TOLIMA
"""
import os, sys, json, io, unicodedata, datetime
import requests
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl

def norm(s):
    s = str(s or '').strip().lower()
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def campo(header):
    h = norm(header)
    if h == 'referencia' or h == 'codigo': return 'codigo'
    if h in ('detalle', 'descripcion', 'articulo', 'nombre'): return 'nombre'
    if h == 'marca': return 'marca'
    if h == 'tat': return 'precioTat'
    if h in ('saldo', 'existencia', 'existencias', 'stock'): return 'stock'
    return None

def parse_xlsx(data: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # localizar fila de encabezados (la que contiene REFERENCIA)
    hi = next((i for i, r in enumerate(rows)
               if any(norm(c) == 'referencia' for c in r)), None)
    if hi is None:
        raise RuntimeError('No encontré la columna REFERENCIA en el archivo')
    colmap = {}
    for ci, c in enumerate(rows[hi]):
        f = campo(c)
        if f and f not in colmap.values():
            colmap[ci] = f
    filas = []
    for r in rows[hi + 1:]:
        d = {}
        for ci, f in colmap.items():
            if ci < len(r):
                d[f] = r[ci]
        cod = str(d.get('codigo') or '').strip()
        if not cod:
            continue  # fila de título de categoría o vacía
        fila = {'codigo': cod}
        if d.get('nombre'): fila['nombre'] = str(d['nombre']).strip()
        if d.get('marca'): fila['marca'] = str(d['marca']).strip()
        try:
            if d.get('precioTat') not in (None, ''): fila['precioTat'] = float(d['precioTat'])
        except Exception: pass
        try:
            fila['stock'] = round(float(d.get('stock'))) if d.get('stock') not in (None, '') else 0
        except Exception:
            fila['stock'] = 0
        filas.append(fila)
    return filas

def main():
    sa = json.loads(os.environ['GDRIVE_SA_KEY'])
    folder = os.environ['GDRIVE_FOLDER_ID']
    api = os.environ['API_BASE'].rstrip('/')
    token = os.environ['IMPORT_TOKEN'].strip()
    regiones = [x.strip().upper() for x in os.environ.get('INV_REGIONES', 'QUINDIO,TOLIMA').split(',') if x.strip()]

    creds = service_account.Credentials.from_service_account_info(
        sa, scopes=['https://www.googleapis.com/auth/drive.readonly'])
    drive = build('drive', 'v3', credentials=creds)

    # listar archivos de la carpeta (recientes primero)
    q = f"'{folder}' in parents and trashed = false"
    res = drive.files().list(q=q, orderBy='modifiedTime desc',
                             fields='files(id,name,mimeType,modifiedTime)', pageSize=200).execute()
    archivos = res.get('files', [])
    print(f"{len(archivos)} archivo(s) en la carpeta.")

    errores = 0
    for region in regiones:
        match = next((f for f in archivos if region in norm(f['name']).upper()
                      or norm(region) in norm(f['name'])), None)
        if not match:
            print(f"[{region}] No hay archivo con '{region}' en el nombre. Salto.")
            continue
        print(f"[{region}] Usando: {match['name']} (mod {match['modifiedTime']})")
        # descargar (export si es Google Sheet)
        if match['mimeType'] == 'application/vnd.google-apps.spreadsheet':
            req = drive.files().export_media(fileId=match['id'],
                  mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            req = drive.files().get_media(fileId=match['id'])
        buf = io.BytesIO(); dl = MediaIoBaseDownload(buf, req)
        done = False
        while not done: _, done = dl.next_chunk()
        try:
            filas = parse_xlsx(buf.getvalue())
        except Exception as e:
            print(f"[{region}] Error leyendo el archivo: {e}"); errores += 1; continue
        if not filas:
            print(f"[{region}] El archivo no tiene filas válidas. Salto."); continue

        r = requests.post(f"{api}/importar/inventario-auto",
                          headers={'x-import-token': token, 'content-type': 'application/json'},
                          json={'region': region, 'archivo': match['name'], 'filas': filas}, timeout=180)
        if r.status_code == 200:
            print(f"[{region}] OK -> {r.json()}")
        else:
            print(f"[{region}] FALLO {r.status_code}: {r.text[:300]}"); errores += 1

    if errores:
        sys.exit(1)
    print("Listo.")

if __name__ == '__main__':
    main()
